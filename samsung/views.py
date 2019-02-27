from django.shortcuts import render
from django.views.generic.base import TemplateView
from django.views.generic.edit import FormView
from django.http import HttpResponse, HttpResponseRedirect
from .forms import UploadFileForm, FileFieldForm
from django.urls import reverse
from mysite import settings
from selenium import webdriver
import time
import random
import os
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

# for linux

# Create your views here.
def excel_export(request):
    file_name = 'se_monitoring_output.xlsx'
    
    #excel download
    filepath = os.path.join(settings.MEDIA_ROOT, file_name)
    print('up',filepath)
    filename = os.path.basename(filepath)

    with open(filepath, 'rb') as f:
        response = HttpResponse(f, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        print('check')
        return response
    
    return render(request, 'samsung/excel_export.html')

def handle_uploaded_file(f):
    with open('input.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['upload_file']
            #not csv file
            if not csv_file.name.endswith('.csv'):
                return HttpResponseRedirect(reverse('samsung:samsung_index'))

            file_data = csv_file.read().decode("utf-8")
            lines = file_data.split("\n")
            processing(lines)
            #handle_uploaded_file(request.FILES['upload_file'])
            return HttpResponseRedirect(reverse('samsung:excel_export'))
    else:
        form = UploadFileForm()
    return render(request, 'samsung/samsung_index.html', {'form': form})

def processing(data_lines):
    structured_data = data_extract(data_lines)
    total_data = make_search_url(structured_data)
    #print(total_data)
    result_check(total_data)
    file_name = 'se_monitoring_output.xlsx'
    excel_result_store(file_name, total_data)

    

def data_extract(data_lines):
    country = list()
    i = 0
    data_lines = data_lines[1:]
    for rows in data_lines:
        if ',' not in rows:
            break
        field = rows.split(",")
        country.append('')
        country[i] = dict()
        country[i]['num'] = field[0]
        country[i]['country'] = field[1]
        country[i]['language'] = field[2]
        country[i]['localPageUrl'] = field[3]
        country[i]['globalPageUrl'] = field[4]
        country[i]['geoCode'] = field[5]
        country[i]['lanCode'] = field[6]
        country[i]['keyword'] = field[7]
        country[i]['screenshotFile'] = field[8].rstrip()
        i = i+1
    
    return country

def make_search_url(country_data):

    for each in country_data:
        kw = each['keyword']
        kw = kw.replace(' ','+')
        if kw[-1] == '+':
            kw = kw[:-1] + '%2B'
        searchUrl = 'https://www.google.com/search?hl=' + each['lanCode'] + '&gl=' + each['geoCode'] + '&q=' + kw
        each['searchUrl'] = searchUrl
    return country_data

def result_check(data):
    
    header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'
            }
    
    # data에서 나라별로 data 추출
    for each in data:
        req = Request(each['searchUrl'], headers=header)
        html = urlopen(req)
        randomSleep(2)
        bs0bj = BeautifulSoup(html, "html.parser")
        
        # 광고와 이미지 등을 제외한 웹페이지 검색 결과는 <h3 class="r"> 로 구분 가능
        bs_raw = str(bs0bj).split('<div class="rc">')

        # 검색결과에서 url 확인할 부분만 추출  ask 카테고리가 있어서 span class로 걸러줘야 함
        bs = list()
        bs.append('')
        for i in range(1,len(bs_raw)):
            if 'span class="st"' in bs_raw[i]:
                bs.append(bs_raw[i].split('ping=')[0])
        
        # local ranking check, rank[0]은 검색 결과 이전 dummy data
        for rank in range(1, len(bs)):
            
            if each['localPageUrl'] in bs[rank]:
                each['localRank'] = rank
                localLandingPage = bs[rank].split('href="')[1].split('"')[0].split('galaxy-s9')[1] # galaxy-s9/ 이 최종 url 부분, 범용화 시 이 부분 변경. 
                # local 랜딩페이지 확인, 주소 galaxy-s9/ 이후에 값 확인
                if len(localLandingPage) > 1:
                    # galaxy-s9으로 split 했을 때 뒤에 나오는 내용
                    each['localLanding'] = localLandingPage.replace('/','')
                else:
                    each['localLanding'] = 'highlights'
                break
            
        # 검색결과에 page 없을 경우 
        if 'localRank' not in each:
            each['localRank'] = '-'
            each['localLanding'] = '-'
        print('country', each['country'])
        print('rank', each['localRank'])
        
        # global ranking check
        for rank in range(1, len(bs)):
            if each['globalPageUrl'] in bs[rank]:
                each['globalRank'] = rank
                globalLandingPage = bs[rank].split('href="')[1].split('"')[0].split('galaxy-s9')[1]
                # global 랜딩페이지 확인, 주소 galaxy-s9/ 이후에 값 확인
                # note8의 경우 note8 로 스플릿
                
                if len(globalLandingPage) > 1:
                    # galaxy-s9으로 split 했을 때 뒤에 나오는 내용
                    each['globalLanding'] = globalLandingPage.replace('/','')
                else:
                    each['globalLanding'] = 'highlights'
                break
            
        if 'globalRank' not in each:
            each['globalRank'] = '-'
            each['globalLanding'] = '-'
        
    # 결과는 data dict에 저장


def randomSleep(second):
    rand = (time.time() * random.randint(1,9) ) % 3
    
    if rand < second:
        rand = rand + second
    
    rand = round ( rand, 2)
    time.sleep(rand)

# excel에 데이터 저장
def excel_result_store(file_name, data):
    wb = Workbook()
    dest_filename = file_name
    
    ws1 = wb.active
    ws1.title = 'serp ranking'
    
    ws1.cell(column = 2, row = 2, value = 'No.')
    ws1.cell(column = 3, row = 2, value = 'Country(Sitecode)')
    ws1.cell(column = 4, row = 2, value = 'language')
    ws1.cell(column = 5, row = 2, value = 'Flagship MKT PD (Highlights)')
    ws1.cell(column = 6, row = 2, value = '지역코드')
    ws1.cell(column = 7, row = 2, value = '언어코드')
    ws1.cell(column = 8, row = 2, value = 'KW')
    ws1.cell(column = 9, row = 2, value = '로컬 rank')
    ws1.cell(column = 10, row = 2, value = '로컬 페이지명')
    ws1.cell(column = 11, row = 2, value = '글로벌 rank')
    ws1.cell(column = 12, row = 2, value = '글로벌 페이지명')
    ws1.cell(column = 13, row = 2, value = 'KW 파일명')
    ws1.cell(column = 14, row = 2, value = '검색결과 URL')
    
    row_num = 3
    num = 1
    for each in data:
        ws1.cell(column = 2, row = row_num, value = num )
        ws1.cell(column = 3, row = row_num, value = each['country'] )
        ws1.cell(column = 4, row = row_num, value = each['language'] )
        ws1.cell(column = 5, row = row_num, value = each['localPageUrl'] )
        ws1.cell(column = 6, row = row_num, value = each['geoCode'] )
        ws1.cell(column = 7, row = row_num, value = each['lanCode'] )
        ws1.cell(column = 8, row = row_num, value = each['keyword'] )
        ws1.cell(column = 9, row = row_num, value = each['localRank'] )
        ws1.cell(column = 10, row = row_num, value = each['localLanding'] )
        ws1.cell(column = 11, row = row_num, value = each['globalRank'] )
        ws1.cell(column = 12, row = row_num, value = each['globalLanding'] )
        ws1.cell(column = 13, row = row_num, value = each['screenshotFile'] )
        ws1.cell(column = 14, row = row_num, value = each['searchUrl'] )
        
        row_num = row_num + 1
        num = num + 1
    filepath = os.path.join(settings.BASE_DIR, 'media')
    print(filepath)
    #########################################
    ### linux는 path 가 / 로 구분
    #print(filepath + '/' + file_name)
    #wb.save(filename = filepath + '/' + file_name)

    wb.save(filename = filepath + '/' + file_name)
