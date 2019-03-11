from django.views.generic.base import TemplateView
from django.views.generic import ListView
from django.views.generic import DetailView
from expedia.models import ExpediaKeywords, ExpediaUrls
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render

import os
import time
from openpyxl import load_workbook, Workbook
from urllib import parse
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
from mysite import settings

# Create your views here.
# TemplateView
class ExpediaModelView(TemplateView):
    template_name = 'expedia/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['expedia_list'] = ['ExpediaKeywords', 'ExpediaUrls']
        return context

def process(request):
    header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'
        }

    # power link
    # make_url
    power_url = list()

    kw=request.POST['kw_field'].split('\n')
    
    company = check_list()

    pc_powerlink_ranking,pc_powerlink_ranking_url = pc_powerlink(kw, company)
    #pc_powerlink_ranking
    #pc_powerlink_ranking_url

    pc_website_ranking, pc_website_url = pc_website(kw,company)
    #pc_website_ranking
    #pc_website_url

    pc_naverpost_ranking, pc_naverpost_url = pc_post(kw, company)
    #pc_naverpost_ranking
    #pc_naverpost_url

    pc_naverblog_ranking, pc_naverblog_url = pc_blog(kw, company)
    #pc_naverblog_ranking
    #pc_naverblog_url

    m_powerlink_ranking, m_powerlink_ranking_url = m_powerlink(kw, company)
    #m_powerlink_ranking

    m_website_ranking, m_website_url = m_web(kw, company)
    #m_website_ranking
    #m_website_url

    m_naverpost_ranking, m_naverpost_url = m_post(kw, company)
    #m_naverpost_ranking
    #m_post_url_ranking

    m_naverblog_ranking, m_naverblog_url = m_blog(kw, company)
    #m_naverblog_ranking
    #m_naverblog_url

    f_name = 'expedia_raw_data.xlsx'
    excel_out(f_name, kw, company, pc_powerlink_ranking, pc_powerlink_ranking_url, pc_website_ranking, pc_website_url, pc_naverpost_ranking, pc_naverpost_url, pc_naverblog_ranking, pc_naverblog_url, m_powerlink_ranking, m_powerlink_ranking_url, m_website_ranking, m_website_url, m_naverpost_ranking, m_naverpost_url, m_naverblog_ranking, m_naverblog_url)
    
    # Excel Download
    # 저장되는 위치 바뀌면 이 경로도 바뀌어야 함 
    filepath = os.path.join(settings.MEDIA_ROOT, f_name)
    filename = os.path.basename(filepath)

    with open(filepath, 'rb') as f:
        response = HttpResponse(f, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        return response
    
    #return render(request, 'expedia/results.html', {'keywords':kw})

def pc_powerlink(kw, company):
    header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'
            }
    
    # power link
    # make_url
    power_url = list()
    for i in range(len(kw)):
        power_url.append('https://ad.search.naver.com/search.naver?where=ad&query=' + str(parse.quote(kw[i])) )
    
    # ranking check list
    pc_power_ranking = list()
    print('kw is : ', kw)
    pc_power_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(power_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")
        
        pc_power_ranking.append('')
        pc_power_ranking[i] = dict()
        pc_power_url.append('')
        pc_power_url[i] = dict()
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.find("a",{"class":"url"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.find_all("a", {"class":"url"})
            urls = list()
            for tag_i in range(len(urls_tag)):
                urls.append(urls_tag[tag_i].text)
                urls[tag_i] = urls[tag_i].replace("https://","")
                urls[tag_i] = urls[tag_i].replace("http://","")
                #if urls[tag_i][-1] == "/":
                #    urls[tag_i] = urls[tag_i][:-1]
            
            # 각 회사마다 랭킹 알아내기
                        
            
            for name in company.keys():
                if company[name]['powerlink'] in urls:
                    pc_power_ranking[i][name] = urls.index(company[name]['powerlink']) + 1
                    pc_power_url[i][name] = urls_tag[urls.index(company[name]['powerlink'])].attrs['href']
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            pc_power_ranking.append('')
            pc_power_ranking[i] = dict()
            for name in company.keys():
                pc_power_ranking[i][name] = 'N'
        '''
        
        time.sleep(0.2)
        
    return pc_power_ranking, pc_power_url

def pc_website(kw, company):
    header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'
            }
    
    # post link
    # make_url
    web_url = list()
    for i in range(len(kw)):
        web_url.append('https://search.naver.com/search.naver?where=webkr&sm=tab_jum&query=' + str(parse.quote(kw[i])) )
        
    pc_web_ranking = list()
    pc_web_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(web_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")

        pc_web_ranking.append('')
        pc_web_ranking[i] = dict()
        pc_web_url.append('')
        pc_web_url[i] = dict()
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.find("a",{"class":"txt_url"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.findAll("a",{"class":"txt_url"})           
            urls = list()
            for tag_i in range(len(urls_tag)):
                urls.append(urls_tag[tag_i].attrs['href'])
                urls[tag_i] = urls[tag_i].replace("https://","")
                urls[tag_i] = urls[tag_i].replace("http://","")
                #if urls[tag_i][-1] == "/":
                #    urls[tag_i] = urls[tag_i][:-1]
            
            # 각 회사마다 랭킹 알아내기
            # website의 경우에는 원본 url 뒤에 하위 url이 붙기 때문에, 각각 url을 체크하며 그 안에 메인 도메인이 있는지 확인하는 작업 필요

            for name in company.keys():
                rank = 1
                for each in urls:
                    if company[name]['website'] in each:
                        pc_web_ranking[i][name] = rank
                        pc_web_url[i][name] = each
                        break
                    rank = rank + 1
                    
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            pc_web_ranking.append('')
            pc_web_ranking[i] = dict()
            for name in company.keys():
                pc_web_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return pc_web_ranking , pc_web_url   

def pc_post(kw, company):
    header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'
            }
    
    # post link
    # make_url
    post_url = list()
    for i in range(len(kw)):
        post_url.append('https://m.post.naver.com/search/post.nhn?keyword=' + str(parse.quote(kw[i])) )
        
    pc_post_ranking = list()
    pc_post_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(post_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")
        
        pc_post_ranking.append('')
        pc_post_ranking[i] = dict()
        pc_post_url.append('')
        pc_post_url[i] = dict()
    
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.findAll("div",{"class":"feed_body"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.findAll("div",{"class":"feed_body"})
            urls = list()
            
            for tag_i in range(len(urls_tag)):
                urls.append(str(urls_tag[tag_i].find("a",{"class":"link_end"}).attrs['href']))
            
            # 각 회사마다 랭킹 알아내기

            for name in company.keys():
                rank = 1
                for each in urls:
                    if company[name]['post'] is not None:
                        if company[name]['post'] in each:
                            pc_post_ranking[i][name] = rank
                            pc_post_url[i][name] = 'https://m.post.naver.com/'+each
                            break
                        rank = rank + 1
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            pc_post_ranking.append('')
            pc_post_ranking[i] = dict()
            for name in company.keys():
                pc_post_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return pc_post_ranking, pc_post_url

def pc_blog(kw, company):
    header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.181 Safari/537.36'
            }
    
    # post link
    # make_url
    blog_url = list()
    for i in range(len(kw)):
        blog_url.append('https://search.naver.com/search.naver?where=post&sm=tab_jum&query=' + str(parse.quote(kw[i])) )
        
    pc_blog_ranking = list()
    pc_blog_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(blog_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")
        
        pc_blog_ranking.append('')
        pc_blog_ranking[i] = dict()
        pc_blog_url.append('')
        pc_blog_url[i] = dict()
    
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.findAll("a",{"class":"sh_blog_title"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.findAll("a",{"class":"sh_blog_title"})
            urls = list()
            
            for tag_i in range(len(urls_tag)):
                urls.append(str(urls_tag[tag_i].attrs['href']))
            
            # 각 회사마다 랭킹 알아내기

            for name in company.keys():
                rank = 1
                for each in urls:
                    if company[name]['blog'] is not None:
                        if company[name]['blog'] in each:
                            pc_blog_ranking[i][name] = rank
                            pc_blog_url[i][name] = each
                            break
                        rank = rank + 1
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            pc_post_ranking.append('')
            pc_post_ranking[i] = dict()
            for name in company.keys():
                pc_post_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return pc_blog_ranking, pc_blog_url

def m_powerlink(kw, company):
    # mobile용 헤더 사용, 크롬 모바일
    header = {
            'User-Agent': 'Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36'
            }
    
    # power link
    # make_url
    m_power_url = list()
    for i in range(len(kw)):
        m_power_url.append('https://m.ad.search.naver.com/search.naver?where=m_expd&query=' + str(parse.quote(kw[i])) )
    
    # ranking check list
    m_power_ranking = list()
    m_power_ranking_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(m_power_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")

        m_power_ranking.append('')
        m_power_ranking[i] = dict()
        m_power_ranking_url.append('')
        m_power_ranking_url[i] = dict()
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.find("cite",{"class":"url"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.find_all("cite", {"class":"url"})
            
            urls = list()
            for tag_i in range(len(urls_tag)):
                urls.append(urls_tag[tag_i].a.text)
                urls[tag_i] = urls[tag_i].replace("https://","")
                urls[tag_i] = urls[tag_i].replace("http://","")
                
                # 이 필터링에서 오류 나기도 함. 고쳐야 됨.밑에 회사마다 랭킹과 url 뽑을 때
                # company[name]['m_powerlink'] in urls 로 하니까 검색어마다 검색되는 페이지에 노출되는 url과 정확하게 일치하는 url만 찾음
                # for문 써서 회사별 url이 검색 url 각각마다 그 안에 있는지 체크하는 방식으로 바뀌어야 함
                if urls[tag_i][-1] == "/":
                    urls[tag_i] = urls[tag_i][:-1]
            
            # 각 회사마다 랭킹 알아내기
            
            for name in company.keys():
                if company[name]['m_powerlink'] in urls:
                    m_power_ranking[i][name] = urls.index(company[name]['m_powerlink']) + 1
                    m_power_ranking_url[i][name] = urls_tag[urls.index(company[name]['m_powerlink'])].a.attrs['href']
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            m_power_ranking.append('')
            m_power_ranking[i] = dict()
            for name in company.keys():
                m_power_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return m_power_ranking, m_power_ranking_url

# web 은 코드 다시 봐야함. web 하다가 post 작업했음 
def m_web(kw, company):
    # mobile용 헤더 사용, 크롬 모바일
    header = {
            'User-Agent': 'Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36'
            }
    
    # power link
    # make_url
    m_web_url = list()
    for i in range(len(kw)):
        m_web_url.append('https://m.search.naver.com/search.naver?display=15&page=2&query=' + str(parse.quote(kw[i])) )
    
    # ranking check list
    m_web_ranking = list()
    m_web_ranking_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(m_web_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")

        m_web_ranking.append('')
        m_web_ranking[i] = dict()
        m_web_ranking_url.append('')
        m_web_ranking_url[i] = dict()        

        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.find("a",{"class":"link_url"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.find_all("a", {"class":"link_url"})

            urls = list()
            for tag_i in range(len(urls_tag)):
                urls.append(urls_tag[tag_i].attrs['href'])
                urls[tag_i] = urls[tag_i].replace("https://","")
                urls[tag_i] = urls[tag_i].replace("http://","")
                #if urls[tag_i][-1] == "/":
                #    urls[tag_i] = urls[tag_i][:-1]
            
            # 각 회사마다 랭킹 알아내기
            

            
            for name in company.keys():
                rank = 1
                for each in urls:
                    # company 정보에 포스트 주소가 있을 경우에만 해당 주소로 search
                    if company[name]['website'] is not None:
                        if company[name]['website'] in each:
                            m_web_ranking[i][name] = rank
                            m_web_ranking_url[i][name] = each
                            break
                        rank = rank + 1
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            m_web_ranking.append('')
            m_web_ranking[i] = dict()
            for name in company.keys():
                m_web_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return m_web_ranking, m_web_ranking_url


def m_post(kw, company):
    # mobile용 헤더 사용, 크롬 모바일
    header = {
            'User-Agent': 'Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36'
            }
    
    # power link
    # make_url
    m_post_url = list()
    # 모바일 view 검색 카테고리에서 나오는 post 순위
    for i in range(len(kw)):
        m_post_url.append('https://m.search.naver.com/search.naver?where=m_view&sm=mtb_jum&query=' + str(parse.quote(kw[i])) )
    
    # ranking check list
    m_post_ranking = list()
    m_post_ranking_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(m_post_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")

        m_post_ranking.append('')
        m_post_ranking[i] = dict()
        m_post_ranking_url.append('')
        m_post_ranking_url[i] = dict()
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.find("a",{"class":"thumb_single"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.find_all("a", {"class":"thumb_single"})

            urls = list()
            for tag_i in range(len(urls_tag)):
                urls.append(urls_tag[tag_i].attrs['href'])
                urls[tag_i] = urls[tag_i].replace("https://","")
                urls[tag_i] = urls[tag_i].replace("http://","")
                #if urls[tag_i][-1] == "/":
                #    urls[tag_i] = urls[tag_i][:-1]
            
            # 각 회사마다 랭킹 알아내기
            
            for name in company.keys():
                rank = 1
                for each in urls:
                    # company 정보에 포스트 주소가 있을 경우에만 해당 주소로 search
                    if company[name]['post'] is not None:
                        if company[name]['post'] in each:
                            m_post_ranking[i][name] = rank
                            m_post_ranking_url[i][name] = each
                            break
                        rank = rank + 1
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            m_post_ranking.append('')
            m_post_ranking[i] = dict()
            for name in company.keys():
                m_post_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return m_post_ranking, m_post_ranking_url

def m_blog(kw, company):
    header = {
            'User-Agent': 'Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36'
            }
    
    # post link
    # make_url
    blog_url = list()
    for i in range(len(kw)):
        blog_url.append('https://m.search.naver.com/search.naver?where=m_view&sm=mtb_jum&query=' + str(parse.quote(kw[i])) )
        
    m_blog_ranking = list()
    m_blog_url = list()
    
    # 검색어마다 회사별 순위 체크
    for i in range(len(kw)):
        print(i)
        req = Request(blog_url[i], headers = header)
        html = urlopen(req)
        bs0bj = BeautifulSoup(html, "html.parser")
        
        m_blog_ranking.append('')
        m_blog_ranking[i] = dict()
        m_blog_url.append('')
        m_blog_url[i] = dict()
    
        
        # 해당 검색어 파워 링크 결과 페이지 존재할 경우
        if bs0bj.findAll("a",{"class":"thumb_single"}) is not None:
                  
            # 각 검색 결과마다 나오는 url 확인 및 filtering
            urls_tag = bs0bj.findAll("a",{"class":"total_tit"})
            urls = list()
            
            for tag_i in range(len(urls_tag)):
                urls.append(str(urls_tag[tag_i].attrs['href']))

            # 각 회사마다 랭킹 알아내기

            for name in company.keys():
                rank = 1
                for each in urls:
                    if company[name]['blog'] is not None:
                        if company[name]['blog'] in each:
                            m_blog_ranking[i][name] = rank
                            m_blog_url[i][name] = each
                            break
                        rank = rank + 1
            
        # 파워링크에 검색어가 아예 없을 경우
        '''
        else:
            pc_post_ranking.append('')
            pc_post_ranking[i] = dict()
            for name in company.keys():
                pc_post_ranking[i][name] = 'N'
        '''
        time.sleep(0.2)
        
    return m_blog_ranking, m_blog_url


def check_list():
    company = dict()
    #post의 경우 href 안에서 memberNo를 확인해야함
    
    company['expedia'] = {
            'powerlink':'www.expedia.co.kr',
            'm_powerlink':'www.expedia.co.kr',
            'website':'www.expedia.co.kr',
            'post':'memberNo=1063900',
            'blog':'kr_expedia'
            }
    company['hotels'] = {
            'powerlink':'kr.hotels.com',
            'm_powerlink':'kr.hotels.com/mobile',
            'website':'kr.hotels.com',
            'post':'memberNo=8591454',
            'blog':None
            }
    company['hotelscombined'] = {
            'powerlink':'www.hotelscombined.co.kr',
            'm_powerlink':'www.hotelscombined.co.kr',
            'website':'www.hotelscombined.co.kr',
            'post':'memberNo=3543910',
            'blog':None
            }
    company['agoda'] = {
            'powerlink':'www.agoda.com',
            'm_powerlink':'www.agoda.com/ko-kr',
            'website':'www.agoda.com',
            'post':None,
            'blog':None
            }
    company['booking'] = {
            'powerlink':'www.booking.com',
            'm_powerlink':'www.booking.com',
            'website':'www.booking.com',
            'post':None,
            'blog':None
            }
    
    return company


def excel_out(f_name, kw, company, pc_powerlink_ranking, pc_powerlink_ranking_url, pc_website_ranking, pc_website_url, pc_naverpost_ranking, pc_naverpost_url, pc_naverblog_ranking, pc_naverblog_url, m_powerlink_ranking, m_powerlink_ranking_url, m_website_ranking, m_website_url, m_naverpost_ranking, m_naverpost_url, m_naverblog_ranking, m_naverblog_url):
    wb = Workbook()
    print('kws are ',kw)
    ws = list()
    ws_count = 0
    for com_name in sorted(company):
        if ws_count == 0:
            ws.append('')
            ws[0] = wb.active
            ws[0].title = com_name
        else:
            ws.append('')
            ws[ws_count] = wb.create_sheet(title=com_name)
        
        # column name on top
        ws[ws_count].cell(column = 3, row = 2, value = "pc파워링크 rank" )
        ws[ws_count].cell(column = 4, row = 2, value = "pc웹사이트 rank" )
        ws[ws_count].cell(column = 5, row = 2, value = "pc포스트 rank" )
        ws[ws_count].cell(column = 6, row = 2, value = "pc블로그 rank" )
        ws[ws_count].cell(column = 7, row = 2, value = "m파워링크 rank" )
        ws[ws_count].cell(column = 8, row = 2, value = "m웹사이트 rank" )
        ws[ws_count].cell(column = 9, row = 2, value = "m포스트 rank" )
        ws[ws_count].cell(column = 10, row = 2, value = "m블로그 rank" )

        ws[ws_count].cell(column = 13, row = 2, value = "pc파워링크 url" )
        ws[ws_count].cell(column = 14, row = 2, value = "pc웹사이트 url" )
        ws[ws_count].cell(column = 15, row = 2, value = "pc포스트 url" )
        ws[ws_count].cell(column = 16, row = 2, value = "pc블로그 url" )
        ws[ws_count].cell(column = 17, row = 2, value = "m파워링크 url" )
        ws[ws_count].cell(column = 18, row = 2, value = "m웹사이트 url" )
        ws[ws_count].cell(column = 19, row = 2, value = "m포스트 url" )
        ws[ws_count].cell(column = 20, row = 2, value = "m블로그 url" )

        row_num = 3
        for i in range(len(kw)):
            ws[ws_count].cell(column = 2, row = row_num, value = kw[i] )
            if com_name in pc_powerlink_ranking[i]:
                ws[ws_count].cell(column = 3, row = row_num, value = pc_powerlink_ranking[i][com_name] )
            if com_name in pc_website_ranking[i]:
                ws[ws_count].cell(column = 4, row = row_num, value = pc_website_ranking[i][com_name] )
            if com_name in pc_naverpost_ranking[i]:
                ws[ws_count].cell(column = 5, row = row_num, value = pc_naverpost_ranking[i][com_name] )
            if com_name in pc_naverblog_ranking[i]:
                ws[ws_count].cell(column = 6, row = row_num, value = pc_naverblog_ranking[i][com_name] )
            if com_name in m_powerlink_ranking[i]:
                ws[ws_count].cell(column = 7, row = row_num, value = m_powerlink_ranking[i][com_name] )
            if com_name in m_website_ranking[i]:
                ws[ws_count].cell(column = 8, row = row_num, value = m_website_ranking[i][com_name] )
            if com_name in m_naverpost_ranking[i]:
                ws[ws_count].cell(column = 9, row = row_num, value = m_naverpost_ranking[i][com_name] )
            if com_name in m_naverblog_ranking[i]:
                ws[ws_count].cell(column = 10, row = row_num, value = m_naverblog_ranking[i][com_name] )
            
            # URL
            if com_name in pc_powerlink_ranking_url[i]:
                ws[ws_count].cell(column = 13, row = row_num, value = pc_powerlink_ranking_url[i][com_name] )  
            if com_name in pc_website_url[i]:
                ws[ws_count].cell(column = 14, row = row_num, value = pc_website_url[i][com_name] )
            if com_name in pc_naverpost_url[i]:
                ws[ws_count].cell(column = 15, row = row_num, value = pc_naverpost_url[i][com_name] )
            if com_name in pc_naverblog_url[i]:
                ws[ws_count].cell(column = 16, row = row_num, value = pc_naverblog_url[i][com_name] )
            if com_name in m_powerlink_ranking_url[i]:
                ws[ws_count].cell(column = 17, row = row_num, value = m_powerlink_ranking_url[i][com_name] )
            if com_name in m_website_url[i]:
                ws[ws_count].cell(column = 18, row = row_num, value = m_website_url[i][com_name] )
            if com_name in m_naverpost_url[i]:
                ws[ws_count].cell(column = 19, row = row_num, value = m_naverpost_url[i][com_name] )
            if com_name in m_naverblog_url[i]:
                ws[ws_count].cell(column = 20, row = row_num, value = m_naverblog_url[i][com_name] )
                
            row_num += 1
            
        ws_count += 1
    filepath = os.path.join(settings.BASE_DIR, 'media')
    wb.save(filename = filepath + '/' + f_name)


class results(ListView):
    model = ExpediaKeywords

# ListView
class KeywordsList(ListView):
    model = ExpediaKeywords

class UrlsList(ListView):
    model = ExpediaUrls

# DetailView
class KeywordsDetail(DetailView):
    model = ExpediaKeywords

class UrlsDetail(DetailView):
    model = ExpediaUrls

